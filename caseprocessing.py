import tkinter as tk
from tkinter import *
from functools import partial
import sys
import datetime
from tkinter.ttk import Treeview
import urllib
import requests
from requests.auth import HTTPBasicAuth
from bs4 import BeautifulSoup
import lxml
import csv
from openpyxl import load_workbook
import webbrowser

from guifuncs import *
from patternfuncs import *
from auxfuncs import *

        
def getCaseList(first_name, last_name, user_entry, pass_entry):
  caseListData = []
  caseList = []
  resultNumber = 0
  currPage = 1
  first_name_string = first_name.get()
  last_name_string = last_name.get()
  global justice_username 
  justice_username = user_entry.get()
  global justice_password
  justice_password = pass_entry.get()
  JUSTICE_SEARCH_URL = 'https://www.nebraska.gov/justice/name.cgi'
  data = {
      'start': 0,
      'party_name': last_name_string + ', ' + first_name_string,
      'indiv_entity_type': 'individual',
      'case_type': [
          'CR',
          'TR',
       ],
      'submit_hidden': currPage,
      'sort': 'party_name',
      'order': 'asc',
      'year': '',
      }
  resp_count = requests.post(JUSTICE_SEARCH_URL, timeout=180, auth=HTTPBasicAuth(justice_username, justice_password), data=data)
  soup_count = BeautifulSoup(resp_count.content, 'lxml')
  num_results_text = soup_count.find_all('strong')[0].text
  num_results = num_results_text.split(" of ")[1].split(" Results")[0]
  resultsPageOffsets = getOffsets(num_results, 25)
  
  for pageOffset in resultsPageOffsets:
    caseListData = []
    data['start'] = pageOffset
    data['submit_hidden'] = currPage
    currPage = currPage + 1
    resp_loop = requests.post(JUSTICE_SEARCH_URL, timeout=180, auth=HTTPBasicAuth(justice_username, justice_password), data=data)
    soup_loop = BeautifulSoup(resp_loop.content, 'lxml')
    rows = soup_loop.find_all('tr')
    for row in rows:
      cols = row.find_all('td')
      cols = [ele.text.strip() for ele in cols]
      caseListData.append([ele for ele in cols if ele]) # get rid of empty values
    
    for caseListDatum in caseListData:
      if (len(caseListDatum) > 0):
        resultNumber = resultNumber + 1
        name = caseListDatum[0].split(" (\xa0")[0]
        party_type = caseListDatum[0].split("\xa0)\nDOB: ")[0].split(" (\xa0")[1].split("\xa0)")[0]
        if "DOB: "in caseListDatum[0]:
          DOB = caseListDatum[0].split("DOB: ")[1]
        else:
          DOB = ""
        county = getCountyName("".join(caseListDatum[1].split())[1:3])
        case_num = "".join(caseListDatum[1].split())
        caption = caseListDatum[2]
        judge = caseListDatum[3]
        attorney = caseListDatum[4]
        caseListItem = (resultNumber, name, party_type, DOB, county, case_num, caption, judge, attorney)
        caseList.append(caseListItem)

  return caseList


def getCases(casetree, first_name, last_name):
  fullCaseNums = []
  selected_cases = casetree.selection()
  for selected_case in selected_cases:
    case_text = casetree.item(selected_case, "values")
    fullCaseNums.append(case_text[5])
  processCases(first_name.get(), last_name.get(), fullCaseNums)
  

def processCases(first_name, last_name, fullCaseNums):
  JUSTICE_CASE_URL = "https://www.nebraska.gov/justice/case.cgi"
  # open the workbook with openpyxl here
  wb = load_workbook('Case Details.xlsx')
  ws = wb['JusticeData']
  xl_row = 2
  for fullCaseNum in fullCaseNums:
    print("Retrieving " + fullCaseNum)
    data = {
      "case_number": fullCaseNum
      }
    resp_docket = requests.post(JUSTICE_CASE_URL, timeout=180, auth=HTTPBasicAuth(justice_username, justice_password), data=data)
    soup_docket = BeautifulSoup(resp_docket.content, 'lxml')
    docket_blocks = soup_docket.find_all('pre')
    num_pre_blocks = len(docket_blocks)
    offense_info_block = docket_blocks[2]
    tables = soup_docket.find_all('table')
    for table in tables:
      if ("Receipt" in table.find('thead').get_text()):
        ledgerTable = table
    register_of_actions_block = docket_blocks[5]
    
    i = str(xl_row)
    ws['A' + i] = getCaseTitle(docket_blocks[0].get_text())
    ws['B' + i] = getClassification(docket_blocks[0].get_text())
    ws['C' + i] = getCountyFromCaption(docket_blocks[0].get_text())
    ws['D' + i] = getCountyOrDistrict(docket_blocks[0].get_text())
    ws['E' + i] = getCaseNumber(docket_blocks[0].get_text())
    ws['F' + i] = getProsecutor(docket_blocks[0].get_text())
    ws['G' + i] = offense_info_block.get_text() # Offense Information  
    ws['H' + i] = docket_blocks[0].get_text() #case summary  
    ws['I' + i] = fullCaseNum
    ws['J' + i] = getFilingDate(docket_blocks[0].get_text()) 
    ws['K' + i] = getCaseStatus(docket_blocks[0].get_text())
    ws['L' + i] = getStatusDate(docket_blocks[0].get_text())
    ws['M' + i] = getDisposition(docket_blocks[0].get_text())
    ws['N' + i] = getRegisterFinesPaid(docket_blocks[5].get_text())
    ws['O' + i] = getFines(offense_info_block.get_text())
    try:
      ws['P' + i] = getPayments(ledgerTable)
      ws['Q' + i] = getFeesAndFines(ledgerTable)
    except NameError:
      ws['P' + i] = "No ledger Table"
      ws['Q' + i] = "No ledger Table"
      print("No Ledger Table")
    ws['R' + i] = getJailSentence(offense_info_block.get_text())
    ws['S' + i] = getSentenceYears(offense_info_block.get_text())
    ws['T' + i] = getProbationTerminated(register_of_actions_block.get_text()) # probation terminated
    ws['U' + i] = getSetAside(offense_info_block.get_text()) # judgment set aside
    ws['V' + i] = soup_docket.find_all('table')[0].get_text()# payments made to the court

    #parties/attorneys to the case section is not pulled

    # arresting officers information not pulled
    # warrant information not pulled
    # court costs information section not pulled 
    #financial activity section not pulled
    # Costs for Recovery table not pulled. This is not always present
    ws['W' + i] = register_of_actions_block.get_text()
    #probation/community service satisfied? house arrest?
    #
    # todo: rearrange columns to match Pika data entry order
    #

    xl_row = xl_row + 1
  fp = "criminal_cases_for_" + first_name + "_" + last_name + "_generated_on_" + datetime.datetime.now().strftime('%Y-%m-%d-%H-%M') + ".xlsx"
  wb.save(fp)
  webbrowser.open(fp)
  
