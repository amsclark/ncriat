# Created by Alexander Clark of Metatheria, LLC
# Creative Commons CC0 v1.0 Universal Public Domain Dedication. No Rights Reserved
# Version 0.2


import tkinter as tk
from tkinter import *
from functools import partial
import sys
import datetime
from tkinter.ttk import Treeview
import urllib
import requests
from requests.auth import HTTPBasicAuth
from bs4 import BeautifulSoup
import lxml
import csv
from openpyxl import load_workbook
import webbrowser

def constructFilterWindow(root, search_button, retrieve_button, select_all_button, clear_all_button, cred_frame, user_entry_label, user_entry, pass_entry_label, pass_entry, casetree, caselistbox_frame, def_frame, first_entry_label, first_name, last_entry_label, last_name):
  root.title("Nebraska Criminal Record Information Automation Tool | Filter")
  # remove all the elements from the previous view
  cred_frame.grid_remove()
  search_button.grid_remove()
  user_entry_label.grid_remove()
  user_entry.grid_remove()
  pass_entry_label.grid_remove()
  pass_entry.grid_remove()
  def_frame.grid_remove()
  first_entry_label.grid_remove()
  first_name.grid_remove()
  last_entry_label.grid_remove()
  last_name.grid_remove()
  #add new elements for this view
  caselistbox_frame.grid(row=0, column=1, rowspan=5, columnspan=30, padx=10, pady=10)
  casetree.grid(row=0, column=0, sticky="W")
  caseList = getCaseList(first_name, last_name, user_entry, pass_entry)
  for case in caseList:
      casetree.insert('', END, values=case)
  retrieve_button.grid(row=7, column=1, sticky="W", padx=10, pady=10)
  select_all_button.grid(row=7, column=2, sticky="W", padx=10, pady=10)
  clear_all_button.grid(row=7, column=3, sticky="W", padx=10, pady=10)

  

def constructMainWindow(root, search_button, cred_frame, user_entry_label, user_entry, pass_entry_label, pass_entry, def_frame, first_entry_label, first_name, last_entry_label, last_name):
  root.title("Nebraska Criminal Record Information Automation Tool | Login") 
  cred_frame.grid(row=0, column=1, rowspan=5, padx=10, pady=10)
  user_entry_label.grid(row=0, column=0, sticky="W")
  user_entry.grid(row=0, column=1, sticky="W")
  pass_entry_label.grid(row=1, column=0, sticky="W")
  pass_entry.grid(row=1, column=1, sticky="W")

  def_frame.grid(row=0, column=2, rowspan=5, padx=10, pady=10)
  first_entry_label.grid(row=0, column=0, sticky="W")
  first_name.grid(row=0, column=1, sticky="W")
  last_entry_label.grid(row=1, column=0, sticky="W")
  last_name.grid(row=1, column=1, sticky="W")

  search_button.grid(row=7, column=1, sticky="W", padx=10, pady=10)

def getCountyName(lookup_number):
  county_numbers_dict = {"Adams" : "14",
  "Antelope" : "26",
  "Arthur" : "91",
  "Banner" : "85",
  "Blaine" : "86",
  "Boone" : "23",
  "Box Butte" : "65",
  "Boyd" : "63",
  "Brown" : "75",
  "Buffalo" : "09",
  "Burt" : "31",
  "Butler" : "25",
  "Cass" : "20",
  "Cedar" : "13",
  "Chase" : "72",
  "Cherry" : "66",
  "Cheyenne" : "39",
  "Clay" : "30",
  "Colfax" : "43",
  "Cuming" : "24",
  "Custer" : "04",
  "Dakota" : "70",
  "Dawes" : "69",
  "Dawson" : "18",
  "Deuel" : "78",
  "Dixon" : "35",
  "Dodge" : "05",
  "Douglas" : "01",
  "Dundy" : "76",
  "Fillmore" : "34",
  "Franklin" : "50",
  "Frontier" : "60",
  "Furnas" : "38",
  "Gage" : "03",
  "Garden" : "77",
  "Garfield" : "83",
  "Gosper" : "73",
  "Grant" : "92",
  "Greeley" : "62",
  "Hall" : "08",
  "Hamilton" : "28",
  "Harlan" : "51",
  "Hayes" : "79",
  "Hitchcock" : "67",
  "Holt" : "36",
  "Hooker" : "93",
  "Howard" : "49",
  "Jefferson" : "33",
  "Johnson" : "57",
  "Kearney" : "52",
  "Keith" : "68",
  "Keya Paha" : "82",
  "Kimball" : "71",
  "Knox" : "12",
  "Lancaster" : "02",
  "Lincoln" : "15",
  "Logan" : "87",
  "Loup" : "88",
  "Madison" : "07",
  "McPherson" : "90",
  "Merrick" : "46",
  "Morrill" : "64",
  "Nance" : "58",
  "Nemaha" : "44",
  "Nuckolls" : "42",
  "Otoe" : "11",
  "Pawnee" : "54",
  "Perkins" : "74",
  "Phelps" : "37",
  "Pierce" : "40",
  "Platte" : "10",
  "Polk" : "41",
  "Red Willow" : "48",
  "Richardson" : "19",
  "Rock" : "81",
  "Saline" : "22",
  "Sarpy" : "59",
  "Saunders" : "06",
  "Scotts Bluff" : "21",
  "Seward" : "16",
  "Sheridan" : "61",
  "Sherman" : "56",
  "Sioux" : "80",
  "Stanton" : "53",
  "Thayer" : "32",
  "Thomas" : "89",
  "Thurston" : "55",
  "Valley" : "47",
  "Washington" : "29",
  "Wayne" : "27",
  "Webster" : "45",
  "Wheeler" : "84",
  "York" : "17"}
  for key, value in county_numbers_dict.items():
    if lookup_number == value:
      return key  
  return "no such county"

def selectAll(casetree):
    for item in casetree.get_children():
        casetree.selection_add(item)

def clearAll(casetree):
    for item in casetree.get_children():
        casetree.selection_remove(item)
        
def getCaseList(first_name, last_name, user_entry, pass_entry):
  caseListData = []
  caseList = []
  resultNumber = 0
  currPage = 1
  first_name_string = first_name.get()
  last_name_string = last_name.get()
  global justice_username 
  justice_username = user_entry.get()
  global justice_password
  justice_password = pass_entry.get()
  JUSTICE_URL = 'https://www.nebraska.gov/justice/name.cgi'
  data = {
      'start': 0,
      'party_name': last_name_string + ', ' + first_name_string,
      'indiv_entity_type': 'individual',
      'case_type': [
          'CR',
          'TR',
       ],
      'submit_hidden': currPage,
      'sort': 'party_name',
      'order': 'asc',
      'year': '',
      }
  resp_count = requests.post(JUSTICE_URL, timeout=180, auth=HTTPBasicAuth(justice_username, justice_password), data=data)
  soup_count = BeautifulSoup(resp_count.content, 'lxml')
  num_results_text = soup_count.find_all('strong')[0].text
  num_results = num_results_text.split(" of ")[1].split(" Results")[0]
  resultsPageOffsets = getOffsets(num_results, 25)
  
  for pageOffset in resultsPageOffsets:
    caseListData = []
    data['start'] = pageOffset
    data['submit_hidden'] = currPage
    currPage = currPage + 1
    resp_loop = requests.post(JUSTICE_URL, timeout=180, auth=HTTPBasicAuth(justice_username, justice_password), data=data)
    soup_loop = BeautifulSoup(resp_loop.content, 'lxml')
    rows = soup_loop.find_all('tr')
    for row in rows:
      cols = row.find_all('td')
      cols = [ele.text.strip() for ele in cols]
      caseListData.append([ele for ele in cols if ele]) # get rid of empty values
    
    for caseListDatum in caseListData:
      if (len(caseListDatum) > 0):
        resultNumber = resultNumber + 1
        name = caseListDatum[0].split(" (\xa0")[0]
        party_type = caseListDatum[0].split("\xa0)\nDOB: ")[0].split(" (\xa0")[1].split("\xa0)")[0]
        if "DOB: "in caseListDatum[0]:
          DOB = caseListDatum[0].split("DOB: ")[1]
        else:
          DOB = ""
        county = getCountyName("".join(caseListDatum[1].split())[1:3])
        case_num = "".join(caseListDatum[1].split())
        caption = caseListDatum[2]
        judge = caseListDatum[3]
        attorney = caseListDatum[4]
        caseListItem = (resultNumber, name, party_type, DOB, county, case_num, caption, judge, attorney)
        caseList.append(caseListItem)

  return caseList


def getCases(casetree, first_name, last_name):
  print("getting cases")
  fullCaseNums = []
  selected_cases = casetree.selection()
  for selected_case in selected_cases:
    case_text = casetree.item(selected_case, "values")
    fullCaseNums.append(case_text[5])
  processCases(first_name.get(), last_name.get(), fullCaseNums)
  

def processCases(first_name, last_name, fullCaseNums):
  # open the workbook with openpyxl here
  wb = load_workbook('Case Details.xlsx')
  ws = wb['JusticeData']
  xl_row = 2
  for fullCaseNum in fullCaseNums:
    i = str(xl_row)
    ws['A' + i] = fullCaseNum
    xl_row = xl_row + 1
  fp = "criminal_cases_for_" + first_name + "_" + last_name + "_generated_on_" + datetime.datetime.now().strftime('%Y-%m-%d-%H-%M') + ".xlsx"
  wb.save(fp)
  webbrowser.open(fp)
  


def getOffsets(numResults, limitPerPage):
  offsetsList = []
  number_of_offsets = int(int(numResults)/limitPerPage) + 1
  for x in range(number_of_offsets):
    offsetsList.append(x * limitPerPage)
  return offsetsList

def getCaseSoup(fullCaseNum):
  print("getting case " + fullCaseNum)
  #code to pull the docket page into a beautifulSoup goes here
  # should hopefully be able to access justice_username and justice_password as globals here
  return "beautiful soup for " + fullCaseNum + " goes here"

def extractInfo(caseSoup):
  print("extracting info from")
  return "extracted info"

def appendCaseToSpreadsheet(extractedInfo):
  print("creating spreadsheet")




root = tk.Tk()

#buttons to press
retrieve_button = tk.Button(text="Retrieve Cases", command=lambda: getCases(casetree, first_name, last_name))
select_all_button = tk.Button(text="Select All", command=lambda: selectAll(casetree))
clear_all_button = tk.Button(text="Clear Selection", command=lambda: clearAll(casetree))
search_button = tk.Button(text="Search Cases", command=lambda: constructFilterWindow(root, search_button, retrieve_button, select_all_button, clear_all_button, cred_frame, user_entry_label, user_entry, pass_entry_label, pass_entry, casetree, caselistbox_frame, def_frame, first_entry_label, first_name, last_entry_label, last_name))

# credentials
cred_frame = LabelFrame(root, text="Justice Login Credentials", padx=5, pady=5, relief=RIDGE)
user_entry_label = Label(cred_frame, text="Username")
user_entry=Entry(cred_frame)
pass_entry_label = Label(cred_frame, text="Password")
pass_entry=Entry(cred_frame)
pass_entry.config(show="*")

# defendant name
def_frame = LabelFrame(root, text="Defendant Name", padx=5, pady=5, relief=RIDGE)
first_entry_label = Label(def_frame, text="First Name")
first_name=Entry(def_frame)
last_entry_label = Label(def_frame, text="Last Name")
last_name=Entry(def_frame)


# case list
caselistbox_frame = LabelFrame(root, text = "Cases list", padx=5, pady=5, relief=RIDGE)


columns = ("resultNumber", "party_name", "party_type", "DOB", "county", "case_num", "caption", "judge", "attorney")
casetree = Treeview(caselistbox_frame, columns=columns, show='headings')
casetree.heading('resultNumber', text="#")
casetree.heading('party_name', text="Party Name")
casetree.heading('party_type', text="Party Type")
casetree.heading('DOB', text="DOB")
casetree.heading('county', text="County")
casetree.heading('case_num', text="Case Number")
casetree.heading('caption', text="Caption")
casetree.heading('attorney', text="Judge")
casetree.heading('attorney', text="Attorney")

constructMainWindow(root, search_button, cred_frame, user_entry_label, user_entry, pass_entry_label, pass_entry, def_frame, first_entry_label, first_name, last_entry_label, last_name)

root.mainloop()
